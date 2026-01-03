# tay_weather_bot.py
#
# Tay Township Weather Bot
# - Pulls Environment Canada alerts from regional ATOM feed (source of truth)
# - Writes RSS feed: tay-weather.xml
# - Posts to X automatically (OAuth 2.0 refresh token)
# - Uploads media to X via OAuth 1.0a user context (required for media upload)
# - Posts to Facebook Page automatically (Page access token), supports photo carousels
# - Supports cooldowns + dedupe
#
# REQUIRED GitHub Secrets (X OAuth 2.0 posting):
#   X_CLIENT_ID
#   X_CLIENT_SECRET
#   X_REFRESH_TOKEN
#
# REQUIRED GitHub Secrets (X media upload via OAuth 1.0a user context):
#   X_API_KEY
#   X_API_SECRET
#   X_ACCESS_TOKEN
#   X_ACCESS_TOKEN_SECRET
#
# REQUIRED GitHub Secrets (Facebook Page posting):
#   FB_PAGE_ID
#   FB_PAGE_ACCESS_TOKEN
#
# OPTIONAL workflow env vars:
#   ENABLE_X_POSTING=true|false
#   ENABLE_FB_POSTING=true|false
#   TEST_TWEET=true
#   ALERT_FEED_URL=<ATOM feed url>
#   TAY_COORDS_URL=<coords link>
#   CR29_NORTH_IMAGE_URL=<direct image url OR https://511on.ca/map/Cctv/<id>>
#   CR29_SOUTH_IMAGE_URL=<direct image url OR https://511on.ca/map/Cctv/<id>>
#   ON511_CAMERA_KEYWORD=<default: CR-29>
#
import base64
import datetime as dt
import email.utils
import hashlib
import json
import mimetypes
import os
import random
import re
import time
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional, Tuple

import requests
from requests_oauthlib import OAuth1
import facebook_poster as fb  # FB rate-limited safe posting
import facebook_poster as fb
fb.load_image_bytes = load_image_bytes  # reuse your existing function

# Optional: Excel-backed content configuration
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # type: ignore

# Optional: Google Sheets/Drive (private online config + media)
try:
    from google.oauth2 import service_account  # type: ignore
    from googleapiclient.discovery import build  # type: ignore
    from googleapiclient.http import MediaIoBaseDownload  # type: ignore
except Exception:
    service_account = None  # type: ignore
    build = None  # type: ignore
    MediaIoBaseDownload = None  # type: ignore



# ----------------------------
# Feature toggles
# ----------------------------
ENABLE_X_POSTING = os.getenv("ENABLE_X_POSTING", "false").lower() == "true"
ENABLE_FB_POSTING = os.getenv("ENABLE_FB_POSTING", "false").lower() == "true"
TEST_TWEET = os.getenv("TEST_TWEET", "false").lower() == "true"


# ----------------------------
# Paths
# ----------------------------
STATE_PATH = "state.json"
RSS_PATH = "tay-weather.xml"
ROTATED_X_REFRESH_TOKEN_PATH = "x_refresh_token_rotated.txt"

# Content configuration source:
# - "google": load from Google Sheet
# - "xlsx": load from local Excel file in repo
# - "auto": prefer Google if secrets exist, else fall back to xlsx
CONTENT_CONFIG_XLSX = os.getenv("CONTENT_CONFIG_XLSX", "content_config.xlsx").strip() or "content_config.xlsx"
CONTENT_CONFIG_SOURCE = os.getenv("CONTENT_CONFIG_SOURCE", "auto").strip().lower()
if CONTENT_CONFIG_SOURCE not in {"auto", "google", "xlsx"}:
    CONTENT_CONFIG_SOURCE = "auto"

# Google private config + media (from GitHub Secrets)
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "").strip()
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "").strip()
GOOGLE_SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()

    
# Optional: Telegram approval (GO/NO-GO) gate
ENABLE_TELEGRAM_APPROVAL = os.getenv("ENABLE_TELEGRAM_APPROVAL", "false").lower() == "true"
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()

try:
    # Max seconds to spend polling Telegram per run (keep low for GitHub Actions)
    TELEGRAM_POLL_SECONDS = int(os.getenv("TELEGRAM_POLL_SECONDS", "6"))
except Exception:
    TELEGRAM_POLL_SECONDS = 6

try:
    # How long a pending approval is kept before it's dropped (hours)
    TELEGRAM_APPROVAL_TTL_HOURS = int(os.getenv("TELEGRAM_APPROVAL_TTL_HOURS", "72"))
except Exception:
    TELEGRAM_APPROVAL_TTL_HOURS = 72

USER_AGENT = "tay-weather-rss-bot/1.1"

# Public “more info” URL (Tay coords format)
TAY_COORDS_URL = os.getenv(
    "TAY_COORDS_URL",
    "https://weather.gc.ca/en/location/index.html?coords=44.751,-79.768",
).strip()
MORE_INFO_URL = TAY_COORDS_URL

ALERT_FEED_URL = os.getenv("ALERT_FEED_URL", "https://weather.gc.ca/rss/battleboard/onrm94_e.xml").strip()
DISPLAY_AREA_NAME = "Tay Township area"

# Ontario 511 cameras API
ON511_CAMERAS_API = "https://511on.ca/api/v2/get/cameras"
ON511_CAMERA_KEYWORD = os.getenv("ON511_CAMERA_KEYWORD", "CR-29").strip() or "CR-29"


# ----------------------------
# Severity emoji
# ----------------------------

def severity_emoji(title: str) -> str:
    """Advisory=🟡, Watch=🟠, Warning=🔴, other=⚪"""
    t = (title or "").lower()
    if "warning" in t:
        return "🔴"
    if "watch" in t:
        return "🟠"
    if "advisory" in t:
        return "🟡"
    return "⚪"


# ----------------------------
# Cooldown policy
# ----------------------------
COOLDOWN_MINUTES = {
    "warning": 60,
    "watch": 120,
    "advisory": 180,
    "statement": 240,
    "alert": 180,
    "allclear": 60,
    "default": 180,
}
GLOBAL_COOLDOWN_MINUTES = 5


# ----------------------------
# Generic helpers
# ----------------------------

def normalize(s: str) -> str:
    if not s:
        return ""
    s = s.lower()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def now_utc() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def safe_int(x: Any, default: int) -> int:
    try:
        return int(x)
    except Exception:
        return default


# ----------------------------
# Alert parsing + content config
# ----------------------------

def _level_and_colour(title: str) -> Tuple[str, str]:
    """Returns (level, colour). colour is a friendly label (yellow/orange/red/grey)."""
    t = (title or "").lower()
    if "warning" in t:
        return "warning", "red"
    if "watch" in t:
        return "watch", "orange"
    if "advisory" in t:
        return "advisory", "yellow"
    if "statement" in t:
        return "statement", "yellow"
    return "alert", "grey"


_TYPE_KEYWORDS = [
    ("rainfall", "rainfall"),
    ("heavy rain", "rainfall"),
    ("wind", "wind"),
    ("thunderstorm", "thunderstorm"),
    ("tornado", "tornado"),
    ("snow squall", "snow"),
    ("snow", "snow"),
    ("blizzard", "snow"),
    ("winter storm", "winter"),
    ("ice storm", "freezing_rain"),
    ("freezing rain", "freezing_rain"),
    ("heat", "heat"),
    ("cold", "cold"),
    ("fog", "fog"),
    ("air quality", "air_quality"),
]


def alert_meta_from_title(title: str) -> Dict[str, str]:
    """Extracts coarse metadata used for care statements + media rules."""
    level, colour = _level_and_colour(title)
    t = normalize(title)

    # Try to extract the phrase immediately preceding the level word
    type_phrase = ""
    m = re.search(r"(.+?)\s+(warning|watch|advisory|statement)\b", t)
    if m:
        type_phrase = m.group(1).strip()
    else:
        type_phrase = t

    type_key = "general"
    for needle, key in _TYPE_KEYWORDS:
        if needle in t:
            type_key = key
            break
    if type_key == "general" and type_phrase:
        type_key = re.sub(r"[^a-z0-9]+", "_", type_phrase).strip("_") or "general"

    return {
        "level": level,
        "colour": colour,
        "type": type_key,
        "type_phrase": type_phrase or "general",
    }


def _matches(rule_val: str, actual: str) -> bool:
    rv = normalize(str(rule_val or ""))
    av = normalize(str(actual or ""))
    return (not rv) or (rv == "*") or (rv == av)


def _weighted_choice(rows: List[Dict[str, Any]], seed: str) -> Optional[Dict[str, Any]]:
    if not rows:
        return None
    # Deterministic selection so reruns don't randomly change the message
    rnd = random.Random(int(hashlib.sha1((seed or "").encode("utf-8")).hexdigest(), 16))

    weights: List[float] = []
    for r in rows:
        w = safe_int(r.get("weight", 1), 1)
        # Prefer more-specific rows over wildcard rows
        specificity = 0
        for k in ("colour", "level", "type"):
            v = normalize(str(r.get(k, "")))
            if v and v != "*":
                specificity += 1
        weights.append(max(0.1, float(w)) * (1.0 + specificity * 1.5))

    total = sum(weights)
    if total <= 0:
        return rows[0]
    pick = rnd.random() * total
    upto = 0.0
    for r, w in zip(rows, weights):
        upto += w
        if upto >= pick:
            return r
    return rows[-1]



def _google_creds():
    """Build Google service account creds from env JSON."""
    if not service_account:
        return None
    if not GOOGLE_SERVICE_ACCOUNT_JSON:
        return None
    try:
        info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]
        return service_account.Credentials.from_service_account_info(info, scopes=scopes)
    except Exception as e:
        print(f"⚠️ Google credentials invalid: {e}")
        return None


def _read_google_sheet_tab(sheet_id: str, tab: str, creds) -> List[List[Any]]:
    """Returns rows for a tab using Sheets API."""
    if not build:
        return []
    try:
        svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
        rng = f"{tab}!A1:Z2000"
        resp = svc.spreadsheets().values().get(spreadsheetId=sheet_id, range=rng).execute()
        return resp.get("values", []) or []
    except Exception as e:
        print(f"⚠️ Could not read Google Sheet tab {tab}: {e}")
        return []


def load_content_config() -> Dict[str, Any]:
    """Loads CareStatements + MediaRules + CustomText from either Google Sheet (preferred) or local Excel."""
    cfg: Dict[str, Any] = {"care": [], "media": [], "custom": []}
    print(
        f"Content config source={CONTENT_CONFIG_SOURCE} | "
        f"sheet_id={'set' if GOOGLE_SHEET_ID else 'missing'} | "
        f"service_account={'set' if GOOGLE_SERVICE_ACCOUNT_JSON else 'missing'}"
    )

    def normalize_header(h: str) -> str:
        return normalize(str(h or ""))

    def rows_to_dicts(rows: List[List[Any]]) -> List[Dict[str, Any]]:
        if not rows:
            return []
        headers = [normalize_header(h) for h in rows[0]]
        out: List[Dict[str, Any]] = []
        for r in rows[1:]:
            r = list(r) + [None] * max(0, len(headers) - len(r))
            if not any(x is not None and str(x).strip() for x in r):
                continue
            d: Dict[str, Any] = {}
            for h, v in zip(headers, r):
                if h:
                    d[h] = v
            enabled = str(d.get("enabled", "true")).strip().lower()
            if enabled in {"false", "0", "no", "n"}:
                continue
            out.append(d)
        return out

    # --- Google Sheet path (private online) ---
    if CONTENT_CONFIG_SOURCE in {"auto", "google"} and GOOGLE_SHEET_ID and GOOGLE_SERVICE_ACCOUNT_JSON:
        creds = _google_creds()
        if creds:
            cfg["care"] = rows_to_dicts(_read_google_sheet_tab(GOOGLE_SHEET_ID, "CareStatements", creds))
            cfg["media"] = rows_to_dicts(_read_google_sheet_tab(GOOGLE_SHEET_ID, "MediaRules", creds))
            cfg["custom"] = rows_to_dicts(_read_google_sheet_tab(GOOGLE_SHEET_ID, "CustomText", creds))
            print(
                f"Loaded Google config rows: "
                f"care={len(cfg['care'])}, "
                f"media={len(cfg['media'])}, "
                f"custom={len(cfg['custom'])}"
            )
            print(f"Drive folder id={'set' if GOOGLE_DRIVE_FOLDER_ID else 'missing'}")

            if cfg["care"] or cfg["media"] or cfg["custom"]:
                return cfg
            if CONTENT_CONFIG_SOURCE == "google":
                print("⚠️ Google sheet returned no data; check tab names and sharing.")
                return cfg

    # --- Local Excel fallback ---
    if not load_workbook:
        return cfg
    path = CONTENT_CONFIG_XLSX
    if not os.path.exists(path):
        return cfg

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"⚠️ content_config.xlsx could not be read: {e}")
        return cfg

    def read_sheet(name: str) -> List[Dict[str, Any]]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows_to_dicts(rows)

    cfg["care"] = read_sheet("CareStatements")
    cfg["media"] = read_sheet("MediaRules")
    cfg["custom"] = read_sheet("CustomText")
    return cfg


_drive_service_cache = None
_drive_file_id_cache: Dict[str, str] = {}


def download_drive_media(name_or_id: str) -> Optional[str]:
    """Download a file from the shared Drive folder into /tmp and return local path.
    - name_or_id: either 'id:<fileId>' or a filename in the GOOGLE_DRIVE_FOLDER_ID folder.
    """
    global _drive_service_cache
    if not (GOOGLE_DRIVE_FOLDER_ID and GOOGLE_SERVICE_ACCOUNT_JSON):
        return None
    creds = _google_creds()
    if not creds or not build or not MediaIoBaseDownload:
        return None

    try:
        if _drive_service_cache is None:
            _drive_service_cache = build("drive", "v3", credentials=creds, cache_discovery=False)
        svc = _drive_service_cache

        key = name_or_id.strip()
        file_id: Optional[str] = None
        if key.lower().startswith("id:"):
            file_id = key[3:].strip()
        else:
            if key in _drive_file_id_cache:
                file_id = _drive_file_id_cache[key]
            else:
                q = f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and name = '{key}' and trashed = false"
                resp = svc.files().list(q=q, fields="files(id,name,mimeType)").execute()
                files = resp.get("files", []) or []
                if not files:
                    print(f"⚠️ Drive media not found in folder: {key}")
                    return None
                file_id = files[0]["id"]
                _drive_file_id_cache[key] = file_id

        out_dir = "/tmp/tay_weather_media"
        os.makedirs(out_dir, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", key) if key and not key.lower().startswith("id:") else (file_id or "drive_file")
        out_path = os.path.join(out_dir, safe_name)
        if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
            return out_path

        request = svc.files().get_media(fileId=file_id)
        with open(out_path, "wb") as f:
            downloader = MediaIoBaseDownload(f, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return out_path
    except Exception as e:
        print(f"⚠️ Could not download Drive media {name_or_id}: {e}")
        return None


def pick_care_statement(cfg: Dict[str, Any], meta: Dict[str, str], seed: str) -> str:
    rows = cfg.get("care") or []
    matched: List[Dict[str, Any]] = []
    for r in rows:
        if _matches(str(r.get("colour", "*")), meta.get("colour", "")) and _matches(str(r.get("level", "*")), meta.get("level", "")) and _matches(str(r.get("type", "*")), meta.get("type", "")):
            matched.append({
                "colour": r.get("colour"),
                "level": r.get("level"),
                "type": r.get("type"),
                "weight": r.get("weight", 1),
                "statement": (r.get("statement") or "").strip(),
            })
    choice = _weighted_choice(matched, seed)
    return (choice.get("statement") or "").strip() if choice else ""


def pick_media_refs(cfg: Dict[str, Any], meta: Dict[str, str], seed: str) -> List[Dict[str, str]]:
    """Returns list of media dicts: {kind, ref}."""
    rows = cfg.get("media") or []
    matched: List[Dict[str, Any]] = []

    for r in rows:
        if _matches(str(r.get("colour", "*")), meta.get("colour", "")) and \
           _matches(str(r.get("level", "*")), meta.get("level", "")) and \
           _matches(str(r.get("type", "*")), meta.get("type", "")):

            kind = normalize(str(r.get("media_kind") or "")) or "local"

            # Support Canadian/Drive columns
            ref = (r.get("media_ref") or "").strip()
            if not ref:
                # If file ID provided, use id:<fileId> form
                fid = (r.get("drive_file_id") or "").strip()
                fname = (r.get("drive_filename") or "").strip()
                if fid:
                    ref = f"id:{fid}"
                elif fname:
                    ref = fname

            if ref:
                matched.append({
                    "colour": r.get("colour"),
                    "level": r.get("level"),
                    "type": r.get("type"),
                    "weight": r.get("weight", 1),
                    "kind": kind,
                    "ref": ref,
                })

    choice = _weighted_choice(matched, seed)
    if not choice:
        return []
    return [{"kind": str(choice.get("kind") or "local"), "ref": str(choice.get("ref") or "")}]


def pick_custom_text(cfg: Dict[str, Any], now: dt.datetime, state: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Returns {mode, message, one_shot} if a custom override is currently enabled."""
    rows = cfg.get("custom") or []
    for r in rows:
        enabled = str(r.get("enabled", "false")).strip().lower()
        if enabled in {"false", "0", "no", "n", ""}:
            continue
        mode = normalize(str(r.get("mode") or "append")) or "append"
        message = (r.get("message") or "").strip()
        if not message:
            continue

        def parse_dt(x: Any) -> Optional[dt.datetime]:
            s = (str(x or "")).strip()
            if not s:
                return None
            try:
                d = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
                return d if d.tzinfo else d.replace(tzinfo=dt.timezone.utc)
            except Exception:
                return None

        start = parse_dt(r.get("starts_utc") or r.get("start_utc"))
        end = parse_dt(r.get("ends_utc") or r.get("end_utc"))

        if start and now < start:
            continue
        if end and now > end:
            continue

        one_shot = str(r.get("one_shot", "false")).strip().lower() in {"true", "1", "yes", "y"}
        if one_shot:
            used = set(state.get("custom_one_shots_used", []) or [])
            mh = text_hash(message)
            if mh in used:
                continue

        return {"mode": mode, "message": message, "one_shot": one_shot}
    return None


def text_hash(s: str) -> str:
    return hashlib.sha1((s or "").encode("utf-8")).hexdigest()


def load_state() -> dict:
    default = {
        "seen_ids": [],
        "posted_guids": [],
        "posted_text_hashes": [],
        "cooldowns": {},
        "global_last_post_ts": 0,
        "telegram_last_update_id": 0,
        "pending_approvals": {},
        "approval_decisions": {},
        "token_to_guid": {},
        "custom_one_shots_used": [],
    }
    if not os.path.exists(STATE_PATH):
        return default

    try:
        raw = open(STATE_PATH, "r", encoding="utf-8").read().strip()
        if not raw:
            return default
        data = json.loads(raw)
        if not isinstance(data, dict):
            return default
    except Exception:
        return default

    data.setdefault("seen_ids", [])
    data.setdefault("posted_guids", [])
    data.setdefault("posted_text_hashes", [])
    data.setdefault("cooldowns", {})
    data.setdefault("global_last_post_ts", 0)
    data.setdefault("telegram_last_update_id", 0)
    data.setdefault("pending_approvals", {})
    data.setdefault("approval_decisions", {})
    data.setdefault("token_to_guid", {})
    data.setdefault("custom_one_shots_used", [])
    return data


def save_state(state: dict) -> None:
    state["seen_ids"] = state.get("seen_ids", [])[-5000:]
    state["posted_guids"] = state.get("posted_guids", [])[-5000:]
    state["posted_text_hashes"] = state.get("posted_text_hashes", [])[-5000:]

    cds = state.get("cooldowns", {})
    if isinstance(cds, dict) and len(cds) > 5000:
        items = sorted(cds.items(), key=lambda kv: kv[1], reverse=True)[:4000]
        state["cooldowns"] = dict(items)

    # Prune telegram approval state to keep state.json small
    if isinstance(state.get("approval_decisions"), dict) and len(state["approval_decisions"]) > 1000:
        items = sorted(
            state["approval_decisions"].items(),
            key=lambda kv: safe_int((kv[1] or {}).get("ts", 0), 0),
            reverse=True,
        )[:600]
        state["approval_decisions"] = dict(items)
    if isinstance(state.get("pending_approvals"), dict) and len(state["pending_approvals"]) > 1000:
        items = sorted(
            state["pending_approvals"].items(),
            key=lambda kv: safe_int((kv[1] or {}).get("created_ts", 0), 0),
            reverse=True,
        )[:600]
        state["pending_approvals"] = dict(items)

    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


# ----------------------------
# ATOM helpers
# ----------------------------
ATOM_NS = {"a": "http://www.w3.org/2005/Atom"}


def _parse_atom_dt(s: str) -> dt.datetime:
    if not s:
        return dt.datetime(1970, 1, 1, tzinfo=dt.timezone.utc)
    return dt.datetime.fromisoformat(s.replace("Z", "+00:00"))


def fetch_atom_entries(
    feed_url: str,
    retries: int = 3,
    timeout: Tuple[int, int] = (5, 20),
) -> List[Dict[str, Any]]:
    """Fetch and parse an ATOM feed. Returns entries newest-first."""
    last_err: Optional[Exception] = None
    for attempt in range(retries):
        try:
            r = requests.get(feed_url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
            r.raise_for_status()
            root = ET.fromstring(r.content)

            entries: List[Dict[str, Any]] = []
            for e in root.findall("a:entry", ATOM_NS):
                title = (e.findtext("a:title", default="", namespaces=ATOM_NS) or "").strip()

                link = ""
                link_el = e.find("a:link[@type='text/html']", ATOM_NS)
                if link_el is None:
                    link_el = e.find("a:link", ATOM_NS)
                if link_el is not None:
                    link = (link_el.get("href") or "").strip()

                updated = (e.findtext("a:updated", default="", namespaces=ATOM_NS) or "").strip()
                published = (e.findtext("a:published", default="", namespaces=ATOM_NS) or "").strip()
                entry_id = (e.findtext("a:id", default="", namespaces=ATOM_NS) or "").strip()
                summary = (e.findtext("a:summary", default="", namespaces=ATOM_NS) or "").strip()

                entries.append(
                    {
                        "id": entry_id,
                        "title": title,
                        "link": link,
                        "updated": updated,
                        "published": published,
                        "summary": summary,
                        "updated_dt": _parse_atom_dt(updated or published),
                    }
                )

            entries.sort(key=lambda x: x["updated_dt"], reverse=True)
            return entries
        except Exception as e:
            last_err = e
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
            raise
    raise last_err if last_err else RuntimeError("Failed to fetch ATOM feed")


def atom_title_for_tay(title: str) -> str:
    if not title:
        return title
    t = title.replace(", Midland - Coldwater - Orr Lake", f" ({DISPLAY_AREA_NAME})")
    t = t.replace("Midland - Coldwater - Orr Lake", DISPLAY_AREA_NAME)
    return t


def atom_entry_guid(entry: Dict[str, Any]) -> str:
    return (entry.get("id") or entry.get("link") or entry.get("title") or "").strip()


def build_social_text_from_atom(
    entry: Dict[str, Any],
    care_statement: str = "",
    custom: Optional[Dict[str, Any]] = None,
) -> str:
    """Builds X/Facebook text.

    If custom.mode is "replace", the custom text replaces the generated copy.
    If custom.mode is "append", the custom text is appended.
    """
    title = atom_title_for_tay((entry.get("title") or "").strip())
    issued = (entry.get("summary") or "").strip()

    sev = severity_emoji(title)

    mode = normalize(str((custom or {}).get("mode") or "append")) if custom else "append"
    custom_msg = (custom or {}).get("message") if custom else ""
    custom_msg = (custom_msg or "").strip()

    if custom_msg and mode == "replace":
        parts = [custom_msg]
    else:
        parts = [f"{sev} {title}"]
        if issued:
            parts.append(issued)
        if care_statement:
            parts.append(care_statement)
        if custom_msg and mode == "append":
            parts.append(custom_msg)

    # Always include a stable link + hashtags unless the custom message already includes them
    if MORE_INFO_URL and not any("more:" in normalize(p) for p in parts):
        parts.append(f"More: {MORE_INFO_URL}")
    tags = "#TayTownship #ONStorm"
    if not any("#taytownship" in normalize(p) for p in parts):
        parts.append(tags)

    text = " | ".join([p for p in parts if str(p).strip()])
    return text if len(text) <= 280 else (text[:277].rstrip() + "…")


def build_rss_description_from_atom(entry: Dict[str, Any]) -> str:
    title = atom_title_for_tay((entry.get("title") or "").strip())
    issued = (entry.get("summary") or "").strip()
    official = (entry.get("link") or "").strip()
    bits = [title]
    if issued:
        bits.append(issued)
    bits.append(f"More info (Tay Township): {MORE_INFO_URL}")
    if official:
        bits.append(f"Official alert details: {official}")
    return "\n".join(bits)


# ----------------------------
# RSS helpers
# ----------------------------

def ensure_rss_exists() -> None:
    if os.path.exists(RSS_PATH):
        return

    rss = ET.Element("rss", version="2.0")
    channel = ET.SubElement(rss, "channel")

    ET.SubElement(channel, "title").text = "Tay Township Weather Statements"
    ET.SubElement(channel, "link").text = "https://weatherpresenter.github.io/tay-weather-rss/"
    ET.SubElement(channel, "description").text = "Automated weather statements and alerts for Tay Township area."
    ET.SubElement(channel, "language").text = "en-ca"

    ET.ElementTree(rss).write(RSS_PATH, encoding="utf-8", xml_declaration=True)


def load_rss_tree() -> Tuple[ET.ElementTree, ET.Element]:
    ensure_rss_exists()
    tree = ET.parse(RSS_PATH)
    root = tree.getroot()
    channel = root.find("channel")
    if channel is None:
        raise RuntimeError("RSS file missing <channel>")
    return tree, channel


def rss_item_exists(channel: ET.Element, guid_text: str) -> bool:
    for item in channel.findall("item"):
        guid = item.find("guid")
        if guid is not None and (guid.text or "").strip() == guid_text:
            return True
    return False


def add_rss_item(channel: ET.Element, title: str, link: str, guid: str, pub_date: str, description: str) -> None:
    item = ET.Element("item")
    ET.SubElement(item, "title").text = title
    ET.SubElement(item, "link").text = link
    g = ET.SubElement(item, "guid")
    g.text = guid
    g.set("isPermaLink", "false")
    ET.SubElement(item, "pubDate").text = pub_date
    ET.SubElement(item, "description").text = description

    insert_index = 0
    for i, child in enumerate(list(channel)):
        if child.tag in {"title", "link", "description", "language", "lastBuildDate"}:
            insert_index = i + 1
    channel.insert(insert_index, item)


def trim_rss_items(channel: ET.Element, max_items: int) -> None:
    items = channel.findall("item")
    if len(items) <= max_items:
        return
    for item in items[max_items:]:
        channel.remove(item)


MAX_RSS_ITEMS = 25


# ----------------------------
# Cooldown logic
# ----------------------------

def group_key_for_cooldown(area_name: str, kind: str) -> str:
    raw = f"{normalize(area_name)}|{normalize(kind)}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def cooldown_allows_post(state: Dict[str, Any], area_name: str, kind: str = "alert") -> Tuple[bool, str]:
    now_ts = int(time.time())

    last_global = safe_int(state.get("global_last_post_ts", 0), 0)
    if last_global and (now_ts - last_global) < (GLOBAL_COOLDOWN_MINUTES * 60):
        return False, f"Global cooldown active ({GLOBAL_COOLDOWN_MINUTES}m)."

    key = group_key_for_cooldown(area_name, kind)
    cooldowns = state.get("cooldowns", {}) if isinstance(state.get("cooldowns"), dict) else {}
    last_ts = safe_int(cooldowns.get(key, 0), 0)

    mins = COOLDOWN_MINUTES.get(kind, COOLDOWN_MINUTES["default"])
    if last_ts and (now_ts - last_ts) < (mins * 60):
        return False, f"Cooldown active for group ({mins}m)."

    return True, "OK"


def mark_posted(state: Dict[str, Any], area_name: str, kind: str = "alert") -> None:
    now_ts = int(time.time())
    key = group_key_for_cooldown(area_name, kind)
    state.setdefault("cooldowns", {})
    state["cooldowns"][key] = now_ts
    state["global_last_post_ts"] = now_ts


# ----------------------------
# Ontario 511 camera resolver
# ----------------------------

_ON511_CAMERAS_CACHE: Optional[List[Dict[str, Any]]] = None


def is_image_url(url: str) -> bool:
    """Does URL respond with Content-Type image/*?"""
    url = (url or "").strip()
    if not url:
        return False

    try:
        r = requests.head(url, allow_redirects=True, headers={"User-Agent": USER_AGENT}, timeout=(5, 15))
        ct = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if r.status_code < 400 and ct.startswith("image/"):
            return True
    except Exception:
        pass

    try:
        r = requests.get(url, allow_redirects=True, headers={"User-Agent": USER_AGENT}, timeout=(5, 20))
        ct = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        return r.status_code < 400 and ct.startswith("image/")
    except Exception:
        return False


def fetch_on511_cameras() -> List[Dict[str, Any]]:
    global _ON511_CAMERAS_CACHE
    if _ON511_CAMERAS_CACHE is not None:
        return _ON511_CAMERAS_CACHE

    r = requests.get(ON511_CAMERAS_API, headers={"User-Agent": USER_AGENT}, timeout=(10, 30))
    r.raise_for_status()
    data = r.json()
    if not isinstance(data, list):
        raise RuntimeError("Unexpected 511 cameras payload (expected list).")

    _ON511_CAMERAS_CACHE = data
    return data


def resolve_on511_views_by_keyword(keyword: str) -> List[Dict[str, Any]]:
    kw = normalize(keyword)
    cams = fetch_on511_cameras()
    out: List[Dict[str, Any]] = []

    for cam in cams:
        name = normalize(str(cam.get("Name") or ""))
        desc = normalize(str(cam.get("Description") or ""))
        if kw and (kw in name or kw in desc):
            views = cam.get("Views") or []
            if isinstance(views, list):
                for v in views:
                    if isinstance(v, dict):
                        out.append(v)

    return out


def pick_north_south_view_urls(views: List[Dict[str, Any]]) -> Tuple[str, str]:
    north = ""
    south = ""

    def normalize_url(u: str) -> str:
        u = (u or "").strip()
        if not u:
            return ""
        if not u.lower().startswith("http"):
            u = "https://511on.ca" + (u if u.startswith("/") else "/" + u)
        return u

    for v in views:
        d = normalize(str(v.get("Description") or ""))
        u = normalize_url(v.get("Url") or "")
        if not u:
            continue
        if ("north" in d or "nb" in d) and not north:
            north = u
        if ("south" in d or "sb" in d) and not south:
            south = u

    if not north or not south:
        urls: List[str] = []
        for v in views:
            u = normalize_url(v.get("Url") or "")
            if u:
                urls.append(u)
        if not north and len(urls) >= 1:
            north = urls[0]
        if not south and len(urls) >= 2:
            south = urls[1]

    return north, south


def resolve_cr29_image_urls() -> List[str]:
    """Resolve up to two image URLs (north, south) with fallbacks."""
    north_env = (os.getenv("CR29_NORTH_IMAGE_URL") or "").strip()
    south_env = (os.getenv("CR29_SOUTH_IMAGE_URL") or "").strip()

    urls: List[str] = []

    for u in [north_env, south_env]:
        if u and is_image_url(u) and u not in urls:
            urls.append(u)

    if len(urls) >= 2:
        return urls[:2]

    try:
        views = resolve_on511_views_by_keyword(ON511_CAMERA_KEYWORD)
        north_api, south_api = pick_north_south_view_urls(views)
        for u in [north_api, south_api]:
            if u and is_image_url(u) and u not in urls:
                urls.append(u)
    except Exception as e:
        print(f"⚠️ 511 camera API resolver skipped: {e}")

    return urls[:2]


# ----------------------------
# X OAuth2 (posting) helpers
# ----------------------------

def write_rotated_refresh_token(new_refresh: str) -> None:
    new_refresh = (new_refresh or "").strip()
    if not new_refresh:
        return
    with open(ROTATED_X_REFRESH_TOKEN_PATH, "w", encoding="utf-8") as f:
        f.write(new_refresh)


def get_oauth2_access_token() -> str:
    client_id = os.getenv("X_CLIENT_ID", "").strip()
    client_secret = os.getenv("X_CLIENT_SECRET", "").strip()
    refresh_token = os.getenv("X_REFRESH_TOKEN", "").strip()

    missing = [k for k, v in [
        ("X_CLIENT_ID", client_id),
        ("X_CLIENT_SECRET", client_secret),
        ("X_REFRESH_TOKEN", refresh_token),
    ] if not v]
    if missing:
        raise RuntimeError(f"Missing required X env vars: {', '.join(missing)}")

    basic = base64.b64encode(f"{client_id}:{client_secret}".encode("utf-8")).decode("ascii")
    headers = {
        "Authorization": f"Basic {basic}",
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": USER_AGENT,
    }

    r = requests.post(
        "https://api.x.com/2/oauth2/token",
        headers=headers,
        data={"grant_type": "refresh_token", "refresh_token": refresh_token},
        timeout=30,
    )
    print("X token refresh status:", r.status_code)
    r.raise_for_status()

    payload = r.json()
    access = payload.get("access_token")
    if not access:
        raise RuntimeError("No access_token returned during refresh.")

    new_refresh = payload.get("refresh_token")
    if new_refresh and new_refresh != refresh_token:
        print("⚠️ X refresh token rotated. Workflow will update the repo secret.")
        write_rotated_refresh_token(new_refresh)

    return access


# ----------------------------
# X media upload helpers (OAuth 1.0a)
# ----------------------------

def load_image_bytes(image_ref: str) -> Tuple[bytes, str]:
    """Loads an image from a URL or a local file path.

    - URL: http(s)://...
    - Local: relative path in the repo (e.g., media/wind.png)
    """
    image_ref = (image_ref or "").strip()
    if not image_ref:
        raise RuntimeError("No image reference provided")

    if re.match(r"^https?://", image_ref, flags=re.IGNORECASE):
        r = requests.get(image_ref, headers={"User-Agent": USER_AGENT}, timeout=(10, 30), allow_redirects=True)
        r.raise_for_status()

        content_type = (r.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if not content_type.startswith("image/"):
            raise RuntimeError(f"URL did not return an image. Content-Type={content_type}")
        return r.content, content_type

    # Local file
    local_path = image_ref
    if not os.path.isabs(local_path):
        local_path = os.path.join(os.getcwd(), local_path)
    if not os.path.exists(local_path):
        raise RuntimeError(f"Local image not found: {image_ref}")

    data = open(local_path, "rb").read()
    ct, _ = mimetypes.guess_type(local_path)
    ct = (ct or "image/png").lower()
    if not ct.startswith("image/"):
        ct = "image/png"
    return data, ct


def x_upload_media(image_ref: str) -> str:
    api_key = os.getenv("X_API_KEY", "").strip()
    api_secret = os.getenv("X_API_SECRET", "").strip()
    access_token = os.getenv("X_ACCESS_TOKEN", "").strip()
    access_secret = os.getenv("X_ACCESS_TOKEN_SECRET", "").strip()

    missing = [k for k, v in [
        ("X_API_KEY", api_key),
        ("X_API_SECRET", api_secret),
        ("X_ACCESS_TOKEN", access_token),
        ("X_ACCESS_TOKEN_SECRET", access_secret),
    ] if not v]
    if missing:
        raise RuntimeError(f"Missing required X OAuth1 env vars: {', '.join(missing)}")

    img_bytes, mime_type = load_image_bytes(image_ref)

    auth = OAuth1(api_key, api_secret, access_token, access_secret)
    upload_url = "https://upload.twitter.com/1.1/media/upload.json"

    files = {"media": ("image", img_bytes, mime_type)}
    r = requests.post(upload_url, auth=auth, files=files, timeout=60)

    print("X media upload status:", r.status_code)
    if r.status_code >= 400:
        raise RuntimeError(f"X media upload failed {r.status_code}")

    j = r.json()
    media_id = j.get("media_id_string") or (str(j.get("media_id")) if j.get("media_id") else "")
    if not media_id:
        raise RuntimeError("X media upload succeeded but no media_id returned")

    return media_id


def post_to_x(text: str, image_urls: Optional[List[str]] = None) -> Dict[str, Any]:
    url = "https://api.x.com/2/tweets"
    access_token = get_oauth2_access_token()

    payload: Dict[str, Any] = {"text": text}

    image_urls = [u for u in (image_urls or []) if (u or "").strip()]
    if image_urls:
        media_ids: List[str] = []
        for u in image_urls[:4]:
            try:
                media_ids.append(x_upload_media(u))
            except Exception as e:
                print(f"⚠️ X media skipped for one image: {e}")
        if media_ids:
            payload["media"] = {"media_ids": media_ids}

    r = requests.post(
        url,
        json=payload,
        headers={
            "Authorization": f"Bearer {access_token}",
            "User-Agent": USER_AGENT,
            "Content-Type": "application/json",
        },
        timeout=20,
    )

    print("X POST /2/tweets status:", r.status_code)

    if r.status_code >= 400:
        detail = ""
        try:
            j = r.json()
            detail = (j.get("detail") or "").lower()
        except Exception:
            pass

        if r.status_code == 403 and "duplicate" in detail:
            raise RuntimeError("X_DUPLICATE_TWEET")

        raise RuntimeError(f"X post failed {r.status_code}")

    return r.json()


# ----------------------------
# Facebook Page posting helpers
# ----------------------------

def _fb_debug_response(r: requests.Response, label: str) -> None:
    """Print useful debug info for Facebook Graph API failures."""
    try:
        print(f"{label} status:", r.status_code)
    except Exception:
        pass

    if r.status_code < 400:
        return

    # Raw body (most useful in Actions logs)
    try:
        print("FB error body:", r.text)
    except Exception:
        pass

    # Structured error (when JSON)
    try:
        j = r.json() or {}
        err = j.get("error") or {}
        if err:
            print(
                "FB error parsed:",
                {
                    "message": err.get("message"),
                    "type": err.get("type"),
                    "code": err.get("code"),
                    "error_subcode": err.get("error_subcode"),
                    "fbtrace_id": err.get("fbtrace_id"),
                },
            )
    except Exception:
        pass


def post_to_facebook_page(message: str) -> Dict[str, Any]:
    page_id = os.getenv("FB_PAGE_ID", "").strip()
    page_token = os.getenv("FB_PAGE_ACCESS_TOKEN", "").strip()
    if not page_id or not page_token:
        raise RuntimeError("Missing FB_PAGE_ID or FB_PAGE_ACCESS_TOKEN")

    url = f"https://graph.facebook.com/v24.0/{page_id}/feed"
    r = requests.post(
        url,
        data={"message": message, "access_token": page_token},
        headers={"User-Agent": USER_AGENT},
        timeout=30,
    )

    _fb_debug_response(r, "FB POST /feed")
    if r.status_code >= 400:
        raise RuntimeError(f"Facebook feed post failed {r.status_code}")
    return r.json()


def post_photo_to_facebook_page(caption: str, image_ref: str) -> Dict[str, Any]:
    page_id = os.getenv("FB_PAGE_ID", "").strip()
    page_token = os.getenv("FB_PAGE_ACCESS_TOKEN", "").strip()
    if not page_id or not page_token:
        raise RuntimeError("Missing FB_PAGE_ID or FB_PAGE_ACCESS_TOKEN")
    if not image_ref:
        raise RuntimeError("Missing image_ref for FB photo post")

    url = f"https://graph.facebook.com/v24.0/{page_id}/photos"

    if re.match(r"^https?://", image_ref, flags=re.IGNORECASE):
        r = requests.post(
            url,
            data={
                "url": image_ref,
                "caption": caption,
                "access_token": page_token,
            },
            headers={"User-Agent": USER_AGENT},
            timeout=30,
        )
    else:
        img_bytes, mime_type = load_image_bytes(image_ref)
        r = requests.post(
            url,
            data={"caption": caption, "access_token": page_token},
            files={"source": ("image", img_bytes, mime_type)},
            headers={"User-Agent": USER_AGENT},
            timeout=30,
        )

    _fb_debug_response(r, "FB POST /photos")
    if r.status_code >= 400:
        raise RuntimeError(f"Facebook photo post failed {r.status_code}")
    return r.json()


def safe_post_facebook_with_limits(
    caption: str,
    image_urls: List[str],
    *,
    has_new_social_event: bool,
    state_path: str = "state.json",
    cooldown_seconds: int = 5400,   # 90 minutes (safe)
    block_seconds: int = 21600,     # 6 hours
) -> Dict[str, Any]:
    """
    - If no new social event: skip
    - If within cooldown: skip
    - If FB blocked us recently: skip
    - Try carousel -> single photo -> text
    - If FB rate-limits (code 368): set blocked_until and skip (do NOT crash job)
    """
    state = load_state(state_path)
    now = dt.datetime.now(dt.timezone.utc)

    def iso(t: dt.datetime) -> str:
        return t.isoformat().replace("+00:00", "Z")

    def parse_iso(s: str) -> Optional[dt.datetime]:
        try:
            if s.endswith("Z"):
                s = s[:-1] + "+00:00"
            return dt.datetime.fromisoformat(s)
        except Exception:
            return None

    if not has_new_social_event:
        print("FB: skip (no new event)")
        return {"skipped": True, "reason": "no_new_event"}

    blocked_until = parse_iso(str(state.get("fb_blocked_until", ""))) if state.get("fb_blocked_until") else None
    if blocked_until and now < blocked_until:
        print(f"FB: skip (blocked until {state['fb_blocked_until']})")
        return {"skipped": True, "reason": "blocked", "blocked_until": state["fb_blocked_until"]}

    last_ok = parse_iso(str(state.get("fb_last_posted_at", ""))) if state.get("fb_last_posted_at") else None
    if last_ok and (now - last_ok).total_seconds() < cooldown_seconds:
        print(f"FB: skip (cooldown {cooldown_seconds}s)")
        return {"skipped": True, "reason": "cooldown"}

    def is_rate_limit(resp: requests.Response) -> bool:
        try:
            j = resp.json()
            err = (j or {}).get("error", {}) or {}
            return (resp.status_code >= 400 and err.get("code") == 368 and str(err.get("error_subcode")) == "1390008")
        except Exception:
            return False

    def record_success():
        state["fb_last_posted_at"] = iso(now)
        state.pop("fb_blocked_until", None)
        save_state(state, state_path)

    def record_block():
        until = now + dt.timedelta(seconds=block_seconds)
        state["fb_blocked_until"] = iso(until)
        save_state(state, state_path)

    # 1) Try carousel
    try:
        result = post_carousel_to_facebook_page(caption, image_urls)
        record_success()
        return {"ok": True, "mode": "carousel", "result": result}
    except Exception as e:
        resp = getattr(e, "response", None)
        if isinstance(resp, requests.Response) and is_rate_limit(resp):
            print("⚠️ FB rate-limited (368). Blocking FB and skipping.")
            record_block()
            return {"skipped": True, "reason": "rate_limited"}
        print(f"⚠️ FB carousel failed: {e}. Trying single photo...")

    # 2) Try single photo
    try:
        if image_urls:
            result = post_photo_to_facebook_page(caption, image_urls[0])
            record_success()
            return {"ok": True, "mode": "single_photo", "result": result}
    except Exception as e:
        resp = getattr(e, "response", None)
        if isinstance(resp, requests.Response) and is_rate_limit(resp):
            print("⚠️ FB rate-limited (368). Blocking FB and skipping.")
            record_block()
            return {"skipped": True, "reason": "rate_limited"}
        print(f"⚠️ FB single photo failed: {e}. Trying text-only...")

    # 3) Try text-only
    try:
        result = post_to_facebook_page(caption)
        record_success()
        return {"ok": True, "mode": "text", "result": result}
    except Exception as e:
        resp = getattr(e, "response", None)
        if isinstance(resp, requests.Response) and is_rate_limit(resp):
            print("⚠️ FB rate-limited (368). Blocking FB and skipping.")
            record_block()
            return {"skipped": True, "reason": "rate_limited"}
        print(f"⚠️ FB text-only failed too: {e}. Skipping FB.")
        return {"skipped": True, "reason": "failed_all_modes"}


# ----------------------------
# Telegram approval gate (optional)
# ----------------------------

def _telegram_enabled() -> bool:
    return ENABLE_TELEGRAM_APPROVAL and bool(TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID)


def telegram_api(method: str, **kwargs: Any) -> Dict[str, Any]:
    base = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
    url = f"{base}/{method}"
    r = requests.post(url, json=kwargs, timeout=15)
    if r.status_code >= 400:
        raise RuntimeError(f"Telegram {method} failed {r.status_code}")
    j = r.json()
    if not j.get("ok"):
        raise RuntimeError(f"Telegram {method} not ok")
    return j


def telegram_send_preview(token: str, text: str) -> None:
    if not _telegram_enabled():
        return
    markup = {
        "inline_keyboard": [
            [
                {"text": "✅ GO", "callback_data": f"GO:{token}"},
                {"text": "❌ NO", "callback_data": f"NO:{token}"},
            ]
        ]
    }
    telegram_api(
        "sendMessage",
        chat_id=TELEGRAM_CHAT_ID,
        text=f"Preview ({token})\n\n{text}",
        reply_markup=markup,
        disable_web_page_preview=False,
    )


def telegram_poll_and_record(state: Dict[str, Any]) -> None:
    """Polls Telegram updates and records GO/NO decisions into state."""
    if not _telegram_enabled():
        return

    last = safe_int(state.get("telegram_last_update_id", 0), 0)
    start = time.time()
    while time.time() - start < max(1, TELEGRAM_POLL_SECONDS):
        try:
            base = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
            params = {"timeout": 0, "allowed_updates": ["callback_query", "message"]}
            if last:
                params["offset"] = last + 1
            r = requests.get(f"{base}/getUpdates", params=params, timeout=15)
            r.raise_for_status()
            j = r.json()
            if not j.get("ok"):
                return
            updates = j.get("result") or []
            if not updates:
                return

            for upd in updates:
                uid = safe_int(upd.get("update_id", 0), 0)
                if uid:
                    last = max(last, uid)

                # Inline button clicks
                cq = upd.get("callback_query") or {}
                data = (cq.get("data") or "").strip()
                if data.startswith("GO:") or data.startswith("NO:"):
                    decision = "go" if data.startswith("GO:") else "no"
                    tok = data.split(":", 1)[1].strip()
                    if tok:
                        state.setdefault("approval_decisions", {})
                        state["approval_decisions"][tok] = {"decision": decision, "ts": int(time.time())}
                    # Acknowledge click
                    try:
                        telegram_api("answerCallbackQuery", callback_query_id=cq.get("id"))
                    except Exception:
                        pass
                    continue

                # Text replies: "GO <token>" or "NO <token>"
                msg = upd.get("message") or {}
                txt = (msg.get("text") or "").strip()
                m = re.match(r"^(go|no)\s+([a-f0-9]{6,16})$", txt, flags=re.IGNORECASE)
                if m:
                    decision = "go" if m.group(1).lower() == "go" else "no"
                    tok = m.group(2).lower()
                    state.setdefault("approval_decisions", {})
                    state["approval_decisions"][tok] = {"decision": decision, "ts": int(time.time())}

            state["telegram_last_update_id"] = last
            return
        except Exception as e:
            print(f"⚠️ Telegram poll failed: {e}")
            return


# ----------------------------
# Main
# ----------------------------

def main() -> None:
    # Clean up any previous rotated token file
    if os.path.exists(ROTATED_X_REFRESH_TOKEN_PATH):
        try:
            os.remove(ROTATED_X_REFRESH_TOKEN_PATH)
        except Exception:
            pass

    camera_image_urls = resolve_cr29_image_urls()

    if TEST_TWEET:
        text = "Test post from Tay weather bot ✅"
        if ENABLE_X_POSTING:
            post_to_x(text, image_urls=camera_image_urls)
        if ENABLE_FB_POSTING:
            state = load_state()
            try:
                fb_result = fb.safe_post_facebook(
                    state,
                    caption=text,
                    image_urls=camera_image_urls,
                    has_new_social_event=True,
                    state_path=STATE_PATH,
                )
                print("FB result:", fb_result)
            except Exception as e:
                print(f"⚠️ Facebook posting encountered an unexpected error; skipping: {e}")
        return

    state = load_state()

    # Pull in any approval decisions first (so an approved post can go out on this run)
    telegram_poll_and_record(state)

    # Load Excel-backed content (care statements, media rules, optional custom text)
    cfg = load_content_config()
    current_custom = pick_custom_text(cfg, now_utc(), state)

    # Prune very old pending approvals
    if isinstance(state.get("pending_approvals"), dict):
        ttl_s = max(1, TELEGRAM_APPROVAL_TTL_HOURS) * 3600
        cutoff = int(time.time()) - ttl_s
        state["pending_approvals"] = {
            k: v for k, v in state.get("pending_approvals", {}).items()
            if safe_int((v or {}).get("created_ts", 0), 0) >= cutoff
        }
    posted = set(state.get("posted_guids", []))
    posted_text_hashes = set(state.get("posted_text_hashes", []))

    tree, channel = load_rss_tree()

    new_rss_items = 0
    social_posted = 0
    social_skipped_cooldown = 0

    try:
        atom_entries = fetch_atom_entries(ALERT_FEED_URL)
    except Exception as e:
        print(f"⚠️ ATOM feed unavailable: {e}")
        print("Exiting cleanly; will retry on next scheduled run.")
        return

    for entry in atom_entries:
        guid = atom_entry_guid(entry)
        if not guid:
            continue

        meta = alert_meta_from_title((entry.get("title") or ""))
        care_statement = pick_care_statement(cfg, meta, seed=guid)
        media_rule = pick_media_refs(cfg, meta, seed=guid)

        chosen_images: List[str] = []
        if media_rule:
            kind = normalize(str(media_rule[0].get("kind") or ""))
            ref = (media_rule[0].get("ref") or "").strip()
            if kind == "cameras":
                chosen_images = camera_image_urls
            elif kind in {"drive", "gdrive", "google_drive"} and ref:
                dl = download_drive_media(ref)
                if dl:
                    chosen_images = [dl]
            elif ref:
                chosen_images = [ref]
        if not chosen_images:
            chosen_images = camera_image_urls

        title = atom_title_for_tay((entry.get("title") or "Weather alert").strip())
        pub_dt = entry.get("updated_dt") or dt.datetime.now(dt.timezone.utc)
        pub_date = email.utils.format_datetime(pub_dt)
        link = MORE_INFO_URL
        description = build_rss_description_from_atom(entry)

        if not rss_item_exists(channel, guid):
            add_rss_item(channel, title=title, link=link, guid=guid, pub_date=pub_date, description=description)
            new_rss_items += 1

        if guid in posted:
            continue

        allowed, reason = cooldown_allows_post(state, DISPLAY_AREA_NAME, kind="alert")
        if not allowed:
            social_skipped_cooldown += 1
            print("Social skipped:", reason)
            continue

        social_text = build_social_text_from_atom(entry, care_statement=care_statement, custom=current_custom)
        h = text_hash(social_text)

        if h in posted_text_hashes:
            print("Social skipped: duplicate text hash already posted")
            posted.add(guid)
            continue


        print("Social preview:", social_text.replace("\n", " "))

        # Optional GO/NO-GO approval via Telegram
        if _telegram_enabled():
            token = hashlib.sha1(guid.encode("utf-8")).hexdigest()[:10]
            state.setdefault("token_to_guid", {})
            state["token_to_guid"][token] = guid

            decisions = state.get("approval_decisions", {}) if isinstance(state.get("approval_decisions"), dict) else {}
            decision = (decisions.get(token) or {}).get("decision")

            if decision == "no":
                print(f"Telegram decision=NO for {token}; skipping social post.")
                posted.add(guid)
                state.get("pending_approvals", {}).pop(token, None)
                continue

            if decision != "go":
                # Not yet approved → send preview once and defer posting
                state.setdefault("pending_approvals", {})
                if token not in state["pending_approvals"]:
                    telegram_send_preview(token, social_text)
                    state["pending_approvals"][token] = {
                        "guid": guid,
                        "created_ts": int(time.time()),
                    }
                else:
                    print(f"Awaiting Telegram approval for {token}")
                continue

        posted_anywhere = False

        if ENABLE_X_POSTING:
            try:
                post_to_x(social_text, image_urls=chosen_images)
                posted_anywhere = True
            except RuntimeError as e:
                if str(e) == "X_DUPLICATE_TWEET":
                    print("X rejected duplicate tweet text; skipping.")
                else:
                    raise

        if ENABLE_FB_POSTING:
            try:
                fb_result = fb.safe_post_facebook(
                    state,
                    caption=social_text,
                    image_urls=chosen_images,
                    has_new_social_event=True,
                    state_path=STATE_PATH,
                )
                print("FB result:", fb_result)
                if fb_result.get("posted"):
                    posted_anywhere = True
            except Exception as e:
                print(f"⚠️ Facebook posting encountered an unexpected error; skipping: {e}")

        if posted_anywhere:
            social_posted += 1
            posted.add(guid)
            posted_text_hashes.add(h)
            mark_posted(state, DISPLAY_AREA_NAME, kind="alert")

            # Mark one-shot custom text as used once it actually goes out
            if current_custom and current_custom.get("one_shot"):
                state.setdefault("custom_one_shots_used", [])
                state["custom_one_shots_used"].append(text_hash(current_custom.get("message") or ""))

            # Clear pending approval state
            if _telegram_enabled():
                tok = hashlib.sha1(guid.encode("utf-8")).hexdigest()[:10]
                if isinstance(state.get("pending_approvals"), dict):
                    state["pending_approvals"].pop(tok, None)
        else:
            print("No social posts sent for this alert.")

    lbd = channel.find("lastBuildDate")
    if lbd is None:
        lbd = ET.SubElement(channel, "lastBuildDate")
    lbd.text = email.utils.format_datetime(now_utc())

    trim_rss_items(channel, MAX_RSS_ITEMS)
    tree.write(RSS_PATH, encoding="utf-8", xml_declaration=True)

    state["posted_guids"] = list(posted)
    state["posted_text_hashes"] = list(posted_text_hashes)
    save_state(state)

    print(
        "Run summary:",
        f"new_rss_items_added={new_rss_items}",
        f"social_posted={social_posted}",
        f"social_skipped_cooldown={social_skipped_cooldown}",
    )


if __name__ == "__main__":
    main()
